import random
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
import matplotlib.pyplot as plt
from prepare import SLOT_EMPTY, SLOT_ONLINE, SLOT_OFFLINE, SLOT_TRAVEL, SLOT_UNWANTED
from mpl_toolkits.mplot3d import Axes3D
from config import *
from deap import base, creator, tools, algorithms
from matplotlib.animation import FuncAnimation, PillowWriter

from data_base import DataBase
from teacher import Teacher
from subject import Subject
from school_class import SchoolClass
from lesson import Lesson
import excel_parser as ep
from prepare import parse_preferences
from builder import TimetableBuilder
from config import hours_filename
from parameters import *

# підготовка даних
print("--- Loading Data ---")
Teacher.reset_registry()
Subject.reset_registry()
SchoolClass.reset_registry()
Lesson.reset_registry()
db = DataBase()
wb_h = openpyxl.load_workbook(hours_filename, data_only=True)
ep.process_table(wb_h.active, db)
print(f"Loaded: {len(db.lesson_list)} lessons, {len(db.teacher_dict)} teachers.")
print(db.lesson_list)
try:
    slots_structure = parse_preferences(db)
    print("Preferences loaded successfully.")
except FileNotFoundError:
    print("ERROR: preferences.xlsx not found! Run init.py first.")
    exit()

# налаштування deap
creator.create("FitnessMulti", base.Fitness, weights=(-1.0, -1.0, -1.0))
creator.create("Individual", list, fitness=creator.FitnessMulti)
toolbox = base.Toolbox()
toolbox.register("indices", random.sample, range(len(db.lesson_list)), len(db.lesson_list))
toolbox.register("individual", tools.initIterate, creator.Individual, toolbox.indices)
toolbox.register("population", tools.initRepeat, list, toolbox.individual)

def eval_genome(individual):
    builder = TimetableBuilder(individual, db, slots_structure)
    return builder.build()

toolbox.register("evaluate", eval_genome)
toolbox.register("mate", tools.cxOrdered)
toolbox.register("mutate", tools.mutShuffleIndexes, indpb=INDPB)
toolbox.register("select", tools.selNSGA2)


def main():
    print(f"--- Starting Evolution (Pop: {POPULATION_SIZE}, Gens: {GENERATIONS}) ---")
    pop = toolbox.population(n=POPULATION_SIZE)
    pareto_front = tools.ParetoFront()
    stats = tools.Statistics(lambda ind: ind.fitness.values)
    stats.register("avg", np.mean, axis=0)
    stats.register("min", np.min, axis=0)

    pop, logbook = algorithms.eaMuPlusLambda(
        pop, toolbox,
        mu=POPULATION_SIZE,
        lambda_=POPULATION_SIZE,
        cxpb=CXPB, mutpb=MUTPB,
        ngen=GENERATIONS,
        stats=stats,
        halloffame=pareto_front,
        verbose=True
    )
    print("\n--- Evolution Finished ---")
    print(f"Pareto Front Size: {len(pareto_front)}")
    for i, ind in enumerate(pareto_front):
        f1, f2, f3 = ind.fitness.values
        print(f"Solution {i + 1}: Student={f1:.0f}, Didactic={f2:.0f}, Teacher={f3:.0f}")
    return pareto_front, logbook


# ВІЗУАЛІЗАЦІЯ
def plot_pareto_3d(front):
    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')

    xs = [ind.fitness.values[0]//1000 for ind in front]  # F1
    ys = [ind.fitness.values[1]//1000 for ind in front]  # F2
    zs = [ind.fitness.values[2]//100 for ind in front]  # F3

    ax.scatter(xs, ys, zs, c=zs, cmap='viridis', s=60, alpha=0.9, edgecolors='k')
    plt.show()


# ЗБЕРЕЖЕННЯ РЕЗУЛЬТАТУ
def save_solution(individual, filename="schedule.xlsx", title_prefix=""):
    builder = TimetableBuilder(individual, db, slots_structure)
    builder.build()
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_font = Font(bold=True)

    fill_online = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    fill_offline = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for c_id, days in builder.schedule.items():
        c_obj = SchoolClass.get_by_id(c_id)
        sheet_title = c_obj.name[:30]
        ws = wb.create_sheet(title=sheet_title)

        days_names = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Нд"]
        for i, d in enumerate(days_names):
            cell = ws.cell(row=1, column=2 + i, value=d)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = fill_header

        for day_idx, lessons in enumerate(days):
            for slot_idx, l_id in enumerate(lessons):
                cell = ws.cell(row=2 + slot_idx, column=2 + day_idx)

                num_cell = ws.cell(row=2 + slot_idx, column=1, value=slot_idx)  # Нумерація з 0
                num_cell.alignment = center_align
                num_cell.font = Font(size=8, color="808080")
                num_cell.border = thin_border
                num_cell.fill = fill_header

                if l_id is not None:
                    lesson = db.lesson_list[l_id]
                    subj = Subject.get_by_id(lesson.subject_id)
                    cell.value = subj.name

                    slot_type = slots_structure[c_id][day_idx][slot_idx]

                    if slot_type == SLOT_OFFLINE:
                        cell.fill = fill_offline
                    else:
                        cell.fill = fill_online
                else:
                    cell.value = "-"
                cell.alignment = center_align
                cell.border = thin_border

        ws.column_dimensions['A'].width = 4
        for col_char in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col_char].width = 16

    wb.save(filename)
    print(f"Saved styled schedule: {filename}")

def plot_convergence(logbook, filename="convergence_plot.png"):
    gen = logbook.select("gen")
    avg_history = logbook.select("avg")
    min_history = logbook.select("min")
    avg_f1, avg_f2, avg_f3 = zip(*avg_history)
    min_f1, min_f2, min_f3 = zip(*min_history)
    fig, (ax1, ax2, ax3) = plt.subplots(3, 1, figsize=(10, 12), sharex=True)
    ax1.plot(gen, avg_f1, label="Average", color='blue', linestyle='--')
    ax1.plot(gen, min_f1, label="Minimum (Best)", color='blue', linewidth=2)
    ax1.set_ylabel('(f1)')
    ax1.set_title('Convergence of Objectives')
    ax1.grid(True)
    ax1.legend()
    ax2.plot(gen, avg_f2, label="Average", color='green', linestyle='--')
    ax2.plot(gen, min_f2, label="Minimum (Best)", color='green', linewidth=2)
    ax2.set_ylabel('(f2)')
    ax2.grid(True)
    ax3.plot(gen, avg_f3, label="Average", color='red', linestyle='--')
    ax3.plot(gen, min_f3, label="Minimum (Best)", color='red', linewidth=2)
    ax3.set_ylabel('(f3)')
    ax3.set_xlabel('Generation')
    ax3.grid(True)
    plt.tight_layout()
    plt.savefig(filename, dpi=300)
    print(f"Convergence plot saved to {filename}")
    plt.show()


def save_teacher_solution(individual, filename="teachers.xlsx"):
    builder = TimetableBuilder(individual, db, slots_structure)
    builder.build()
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_font = Font(bold=True)
    fill_online = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # Home
    fill_offline = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # School
    fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sorted_teachers = sorted(db.teacher_dict.values(), key=lambda t: t.name)

    for teacher in sorted_teachers:
        safe_name = teacher.name.replace(":", "").replace("/", "")[:30]
        ws = wb.create_sheet(title=safe_name)
        days_names = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Нд"]
        for i, d in enumerate(days_names):
            cell = ws.cell(row=1, column=2 + i, value=d)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = fill_header
        timeline = builder.teacher_timeline[teacher.id]

        for day_idx in range(7):  # 7 днів
            for slot_idx in range(lessons_available_per_day):
                num_cell = ws.cell(row=2 + slot_idx, column=1, value=slot_idx)  # Нумерація з 0
                num_cell.alignment = center_align
                num_cell.font = Font(size=8, color="808080")
                num_cell.border = thin_border
                num_cell.fill = fill_header
                cell = ws.cell(row=2 + slot_idx, column=2 + day_idx)
                entry = timeline[day_idx][slot_idx]

                if entry is not None:
                    lesson = db.lesson_list[entry['lesson_id']]
                    subj = Subject.get_by_id(lesson.subject_id)
                    cls_obj = SchoolClass.get_by_id(lesson.school_class_id)
                    cell.value = f"{cls_obj.name}\n{subj.name}"
                    if entry['loc'] == 'SCHOOL':
                        cell.fill = fill_offline
                    else:
                        cell.fill = fill_online
                else:
                    cell.value = "-"

                cell.alignment = center_align
                cell.border = thin_border
        ws.column_dimensions['A'].width = 4
        for col_char in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col_char].width = 18

    wb.save(filename)
    print(f"Saved teacher schedule: {filename}")


def save_rotation_animation(front, filename="pareto_3d_rotation.gif"):
    print(f"--- Rendering 3D Animation ({filename}) ---")
    xs = [ind.fitness.values[0]//1000 for ind in front]  # F1: Students
    ys = [ind.fitness.values[1]//1000 for ind in front]  # F2: Didactic
    zs = [ind.fitness.values[2]//100 for ind in front]  # F3: Teachers
    fig = plt.figure(figsize=(10, 8))
    ax = fig.add_subplot(111, projection='3d')
    ax.scatter(xs, ys, zs, c=zs, cmap='viridis', s=60, alpha=0.9, edgecolors='k')

    def update(angle):
        ax.view_init(elev=30, azim=angle)
        return fig,
    anim = FuncAnimation(fig, update, frames=np.arange(0, 360, 2), interval=50)
    anim.save(filename, writer=PillowWriter(fps=20))

    print(f"Animation saved to {filename}")
    plt.close(fig)


if __name__ == "__main__":
    pareto_front, log = main()

    if len(log) > 0:
        plot_convergence(log)

    if len(pareto_front) > 0:
        plot_pareto_3d(pareto_front)

        try:
            save_rotation_animation(pareto_front, "pareto_front.gif")
        except Exception as e:
            print(f"Animation failed: {e}")
        print("\n--- Saving Best Solutions ---")
        def save_pair(ind, name_suffix):
            f_student = f"schedule_{name_suffix}.xlsx"
            f_teacher = f"teachers_{name_suffix}.xlsx"

            save_solution(ind, f_student)
            save_teacher_solution(ind, f_teacher)
            print(f"--> Saved pair: {f_student} & {f_teacher}")


        # Student Best
        best_student = sorted(pareto_front, key=lambda ind: ind.fitness.values[0])[0]
        print(f"Best Student (f1={best_student.fitness.values[0]})")
        save_pair(best_student, "1_best_student")

        # Didactic Best
        best_didactic = sorted(pareto_front, key=lambda ind: ind.fitness.values[1])[0]
        print(f"Best Didactic (f2={best_didactic.fitness.values[1]})")
        save_pair(best_didactic, "2_best_didactic")

        # Teacher Best
        best_teacher = sorted(pareto_front, key=lambda ind: ind.fitness.values[2])[0]
        print(f"Best Teacher (f3={best_teacher.fitness.values[2]})")
        save_pair(best_teacher, "3_best_teacher")

        # Compromise
        best_compromise = sorted(pareto_front, key=lambda ind: sum(ind.fitness.values))[0]
        print(f"Best Compromise (Sum={sum(best_compromise.fitness.values)})")
        save_pair(best_compromise, "4_compromise")