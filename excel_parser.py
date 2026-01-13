import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple

from teacher import Teacher
from lesson import Lesson
from school_class import SchoolClass
from subject import Subject
from day import Day
from config import *

teacher_name_pattern = r'^[А-ЯІЇЄҐа-яіїєґ]+(?:-[А-ЯІЇЄҐа-яіїєґ]+)* [А-ЯІЇЄҐ]\.[А-ЯІЇЄҐ]\.$'

def get_ws_width(ws):
    return ws.merged_cells.ranges.pop().max_col

def get_subtables_coords(ws):
    TABLE_MIN_COL = 1
    TABLE_MAX_COL = get_ws_width(ws)

    row_separators = set()

    for r in ws.merged_cells.ranges:
        if (
            r.min_row == r.max_row
            and r.min_col == TABLE_MIN_COL
            and r.max_col == TABLE_MAX_COL
        ):
            row_separators.add(r.min_row)

    max_row = ws.max_row

    tables = []
    current_start = None

    def is_table_row(row):
        if row in row_separators:
            return False
        for col in range(TABLE_MIN_COL, TABLE_MAX_COL + 1):
            if ws.cell(row, col).value is not None:
                return True
        return False

    for row in range(1, max_row + 1):
        if is_table_row(row):
            if current_start is None:
                current_start = row
        else:
            if current_start is not None:
                tables.append((current_start, row - 1))
                current_start = None

    if current_start is not None:
        tables.append((current_start, max_row))

    table_coords = [
        {
            "min_row": r1,
            "max_row": r2,
            "min_col": TABLE_MIN_COL,
            "max_col": TABLE_MAX_COL,
        }
        for r1, r2 in tables
    ]
    return table_coords

def get_school_class_dict(ws):
    school_class_dict = {}
    for col in range(2, get_ws_width(ws)+1):
        class_id = ws.cell(row=2, column=col).value
        school_class_dict[class_id] = SchoolClass(class_id)
    return school_class_dict

def process_table(ws, db, starting_column=2):
    db.subtable_coords = get_subtables_coords(ws)
    db.school_class_dict = get_school_class_dict(ws)

    for coords in db.subtable_coords:
        min_row = coords["min_row"]
        max_row = coords["max_row"]
        min_col = coords["min_col"]
        max_col = coords["max_col"]
        subject_name = ws.cell(row=min_row, column=min_col).value
        current_subject_object = Subject(subject_name)
        db.subject_dict[subject_name] = current_subject_object

        for row in range(min_row+2, max_row+1):
            current_teacher_name = ws.cell(row=row, column=1).value
            if not bool(re.fullmatch(teacher_name_pattern, current_teacher_name)):
                continue

            if current_teacher_name in db.teacher_dict.keys():
                current_teacher_object = db.teacher_dict[current_teacher_name]
            else:
                current_teacher_object = Teacher(current_teacher_name)
                db.teacher_dict[current_teacher_name] = current_teacher_object

            for col in range(starting_column, max_col):
                load = ws.cell(row=row, column=col).value
                current_class_object = db.school_class_dict[ws.cell(row=min_row, column=col).value]
                if load is None:
                    continue
                load = str(load)
                whole_lesson_number = int(load.split(".")[0])
                blink_lesson_flag = (load.find(".") != -1)



                for counter in range(whole_lesson_number):
                    db.lesson_list.append(Lesson(current_class_object.id, current_subject_object.id, current_teacher_object.id))
                if blink_lesson_flag:
                    db.lesson_list.append(Lesson(current_class_object.id, current_subject_object.id, current_teacher_object.id, is_blinking=True))

def extract_day_patterns(ws, start_cell, width, height):
    min_row, min_col = coordinate_to_tuple(start_cell)
    max_row = min_row + height - 1
    max_col = min_col + width - 1
    days_list = []
    rows_iterator = ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True
    )
    for row_tuple in rows_iterator:
        row_data = list(row_tuple)
        current_day = Day(row_data)
        days_list.append(current_day)
    return days_list

def create_preferences_workbook(db, filename):
    wb = openpyxl.Workbook()
    ws_slots = wb.active
    ws_slots.title = "slots"
    ws_teachers = wb.create_sheet("teachers")
    ws_subjects = wb.create_sheet("subjects")
    classes = list(db.school_class_dict.keys())
    for i, class_name in enumerate(classes):
        ws_slots.cell(row=2 + i, column=1, value=class_name)
    days = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    for i, day_name in enumerate(days):
        ws_slots.cell(row=1, column=2 + i, value=day_name)
    ref_row, ref_col = coordinate_to_tuple(top_left_cell_of_day_templates)
    for i in range(lessons_available_per_day + 1):
        ws_slots.cell(row=ref_row, column=ref_col + 1 + i, value=i)
    for i in range(1, number_of_day_templates + 1):
        ws_slots.cell(row=ref_row + i, column=ref_col, value=i)
    teachers = list(db.teacher_dict.keys())
    for i, teacher_name in enumerate(teachers):
        ws_teachers.cell(row=2 + i, column=1, value=teacher_name)
    ws_teachers["B1"] = "ПРУ"
    subjects = list(db.subject_dict.keys())
    for i, subject_name in enumerate(subjects):
        ws_subjects.cell(row=2 + i, column=1, value=subject_name)
    ws_subjects["B1"] = "ПРУ"
    wb.save(filename)
    print(f"Файл '{filename}' успішно створено та збережено.")
    return wb
