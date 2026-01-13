import openpyxl
from openpyxl.utils import coordinate_to_tuple
from config import *
from data_base import DataBase
import excel_parser as ep
from teacher import Teacher
from subject import Subject
from school_class import SchoolClass

SLOT_EMPTY = 0
SLOT_ONLINE = 1
SLOT_OFFLINE = 2
SLOT_TRAVEL = 3
SLOT_UNWANTED = 4


CHAR_TO_SLOT = {
    "-": SLOT_EMPTY,
    "O": SLOT_ONLINE,
    "U": SLOT_OFFLINE,
    "T": SLOT_TRAVEL,
    "W": SLOT_UNWANTED,
    " ": SLOT_EMPTY,
    None: SLOT_EMPTY
}


def parse_preferences(db: DataBase):
    wb = openpyxl.load_workbook(preferences_filename, data_only=True)
    _parse_teachers_prefs(wb["teachers"], db)
    _parse_subjects_prefs(wb["subjects"], db)
    slots_matrix = _parse_slots_structure(wb["slots"], db)
    return slots_matrix


def _parse_teachers_prefs(ws, db):
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name: continue
        teacher_obj = db.teacher_dict.get(name)
        if teacher_obj:
            can_offline_val = ws.cell(row=row, column=2).value
            teacher_obj.can_offline = (str(can_offline_val).lower() in ['+', '1', 'true', 'yes'])

            travel_val = ws.cell(row=row, column=3).value
            teacher_obj.travel_time = int(travel_val) if travel_val else 0

            wants_windows = ws.cell(row=row, column=4).value
            teacher_obj.wants_windows = (str(wants_windows).lower() in ['+', '1', 'true', 'yes'])

            max_stack = ws.cell(row=row, column=5).value
            teacher_obj.max_online_lessons_from_underground = int(max_stack) if max_stack else 0


def _parse_subjects_prefs(ws, db):
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name: continue

        subject_obj = db.subject_dict.get(name)
        if subject_obj:
            prio_val = ws.cell(row=row, column=2).value
            subject_obj.priority_offline = (str(prio_val).lower() in ['+', '1', 'true', 'yes'])

            difficulty = ws.cell(row=row, column=3).value
            subject_obj.difficulty = int(difficulty) if difficulty else 0

            preferred_stack = ws.cell(row=row, column=4).value
            subject_obj.preferred_stack = int(preferred_stack) if preferred_stack else 0

            max_stack = ws.cell(row=row, column=5).value
            subject_obj.max_stack = int(max_stack) if max_stack else 0

            max_per_day = ws.cell(row=row, column=5).value
            subject_obj.max_per_day = int(max_per_day) if max_per_day else 0

def _parse_slots_structure(ws, db):
    ref_row, ref_col = coordinate_to_tuple(top_left_cell_of_day_templates)
    patterns = {}

    for i in range(1, number_of_day_templates + 1):
        pat_row = ref_row + i
        pat_types = []
        for slot_idx in range(lessons_available_per_day):
            col_idx = ref_col + 1 + slot_idx
            char_code = ws.cell(row=pat_row, column=col_idx).value
            char_str = str(char_code).upper() if char_code is not None else "-"
            slot_type = CHAR_TO_SLOT.get(char_str, SLOT_EMPTY)
            pat_types.append(slot_type)
        patterns[i] = pat_types
    class_slots = {}

    for row in range(2, ws.max_row + 1):
        class_name = ws.cell(row=row, column=1).value
        if not class_name: break

        c_obj = db.school_class_dict.get(class_name)
        if not c_obj: continue

        class_slots[c_obj.id] = []
        for day_idx in range(7):
            col = 2 + day_idx
            val = ws.cell(row=row, column=col).value
            is_empty_day = (val is None) or (str(val).strip() in ["0", "-", ""])
            if is_empty_day:
                class_slots[c_obj.id].append([SLOT_EMPTY] * lessons_available_per_day)
            elif val in patterns:
                class_slots[c_obj.id].append(patterns[val])
            else:
                print(f"Warning: Unknown pattern '{val}' for class {class_name} at day {day_idx + 1}. Setting as empty.")
                class_slots[c_obj.id].append([SLOT_EMPTY] * lessons_available_per_day)

    return class_slots



if __name__ == "__main__":
    db = DataBase()
    Teacher.reset_registry()
    Subject.reset_registry()
    SchoolClass.reset_registry()
    wb_h = openpyxl.load_workbook(hours_filename, data_only=True)
    ep.process_table(wb_h.active, db)

    print(f"Loaded {len(db.teacher_dict)} teachers.")
    try:
        final_slots = parse_preferences(db)
        print("Preferences parsed successfully.")
        first_class = list(db.school_class_dict.values())[0]
        for i in range(7):
            print(f"Slots for {first_class.name} on {i}: {final_slots[first_class.id][i]}")
    except FileNotFoundError:
        print("Run init.py first to generate preferences.xlsx!")